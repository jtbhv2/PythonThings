import tkinter as tk
import pandas as pd
import requests
import datetime
import re
import ctypes
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table
from openpyxl.styles.fills import GradientFill
from openpyxl.worksheet.page import PageMargins
import openpyxl.styles
import ctypes
import win32com.client
from tkinter import simpledialog
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

def applyFormatting(df, ws, wb, outputPath):
        # Apply formatting##########################################################

    ws.title = 'All Drain Zones'

    # Set the format for the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B7B7B7", end_color="B7B7B7", fill_type="solid")

    # Set column widths based on header content
    column_widths = {
        'A': 12.71,  # Service Request Number
        'B': 14.71,  # Reported Date
        'C': 23.71,  # Location
        'D': 22.71,  # Service Request Type ID
        'E': 61.71,  # Service Request Summary
        'F': 5.71,  # Drain Zone
        'G': 5.71,  # Map Page
        'H': 5.71,  # Map Block
        'I': 13.71,  # Assigned To
        'J': 0.71   # SeeClickFix URL does not matter since it gets hidden later
    }

    # Apply the column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions['J'].hidden = True #Hide J

    # Apply misc formatting

    # Applying a 3-color scale to column B
    col = 'B2:B' + str(ws.max_row)
    for cell in ws['B']:
        cell.number_format = 'm/d/yyyy'  # Apply short date format

    # Create a color scale rule (Red, Yellow, Green)
    color_scale_rule = ColorScaleRule(
        start_type="min", start_color=Color(rgb="FFFF0000"),  # Red
        mid_type="percentile", mid_value=50, mid_color=Color(rgb="FFFFFF00"),  # Yellow
        end_type="max", end_color=Color(rgb="FF00FF00")  # Green
    )

    # Apply the color scale rule
    ws.conditional_formatting.add(col, color_scale_rule)

    # Define formatting styles column D
    flood_prevent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light peach
    flood_prevent_font = Font(color='9C0006')

    repair_reset_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light orange
    repair_reset_font = Font(color='9C5700')

    cavity_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # Light gray

    # Apply formatting rules to column D (assume column index = 4)
    col_d_range = "D2:D" + str(ws.max_row)  # Adjusted to cover all rows

    # Conditional formatting rules col D
    ws.conditional_formatting.add(col_d_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("flood",D2))'], stopIfTrue=False, font=flood_prevent_font, fill=flood_prevent_fill))

    ws.conditional_formatting.add(col_d_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("prevent",D2))'], stopIfTrue=False, font=flood_prevent_font, fill=flood_prevent_fill))

    ws.conditional_formatting.add(col_d_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("repair",D2))'], stopIfTrue=False, font=repair_reset_font, fill=repair_reset_fill))

    ws.conditional_formatting.add(col_d_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("reset",D2))'], stopIfTrue=False, font=repair_reset_font, fill=repair_reset_fill))

    ws.conditional_formatting.add(col_d_range, 
        FormulaRule(formula=['ISNUMBER(SEARCH("cavity",D2))'], stopIfTrue=False, fill=cavity_fill))

    # Apply borders around the data cells
    from openpyxl.styles.borders import Border, Side
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    #Specific cell formats ORDER MATTERS and yes they have to be separate, no idk why

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for cell in ['A1','F1', 'G1', 'H1']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    for cell in ws['D']:
        cell.alignment = Alignment(horizontal='center', vertical='center',shrink_to_fit=True)

    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    for row in range(2, ws.max_row + 1): #Format col a based on j
        if ws[f'J{row}'].value:
            cell = ws[f'A{row}']

            fill = GradientFill(stop=("FF0000", "FFFFFF"))
            cell.fill = fill

    #Wrap the entire thing in a precious table
    table = Table(displayName="DataTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    ws.add_table(table)

    zoneColors = {
        'A': '00FF00',
        'B': 'FFFF00',
        'C': '0000FF',
        'D': 'FFA500'
    }

    #Copy the sheet, then color based on zone 
    for i in range(4):
        newSheet = wb.copy_worksheet(ws)
        zone = chr(65 + i)
        newSheet.title = f'Drain Zone {chr(65 + i)}'

        header_fill = PatternFill(start_color=zoneColors[zone], end_color=zoneColors[zone], fill_type="solid")
        for cell in newSheet[1]:  # Loop through all cells in row 1
            cell.fill = header_fill

        if zone == 'C':
            for cell in newSheet[1]:
                cell.font = Font(color='FFFFFF')

        # Apply conditional formatting AGAIN
        col_b_range = f'B2:B{newSheet.max_row}'  # Adjust the range for column B
        color_scale_rule = ColorScaleRule(
            start_type="min", start_color="FF0000",  # Red
            mid_type="percentile", mid_value=50, mid_color="FFFF00",  # Yellow
            end_type="max", end_color="00FF00"  # Green
        )
        newSheet.conditional_formatting.add(col_b_range, color_scale_rule)

        col_d_range = f'D2:D{newSheet.max_row}'

        newSheet.conditional_formatting.add(
            col_d_range, FormulaRule(formula=['ISNUMBER(SEARCH("flood",D2))'], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        newSheet.conditional_formatting.add(
            col_d_range, FormulaRule(formula=['ISNUMBER(SEARCH("prevent",D2))'], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        newSheet.conditional_formatting.add(
            col_d_range, FormulaRule(formula=['ISNUMBER(SEARCH("repair",D2))'], stopIfTrue=False, fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
        )
        newSheet.conditional_formatting.add(
            col_d_range, FormulaRule(formula=['ISNUMBER(SEARCH("reset",D2))'], stopIfTrue=False, fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
        )
        newSheet.conditional_formatting.add(
            col_d_range, FormulaRule(formula=['ISNUMBER(SEARCH("cavity",D2))'], stopIfTrue=False, fill=PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid"))
        )

    #Pseudo Filter zones
    for i in range (2,6):
        sheet = wb[f'Drain Zone {chr(65 + i - 2)}']
        drainZone = chr(65 + i - 2)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
            if row[0].value != drainZone:
                sheet.row_dimensions[row[0].row].hidden = True

    #Let's get this stupid thing ready to print
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Landscape orientation
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL  # Legal paper size
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.3, footer=0.3)
        ws.oddFooter.right.text = "&\"-,Bold\"&14&KFF0000&D"  # Date in red
        ws.oddFooter.left.text = "&\"-,Bold\"&14&KFF0000&P"  # Page number in the left footer in red


    # Save the workbook with formatting applied
    wb.save(outputPath)
    print(f"Data saved and formatted successfully to {outputPath}")

def incinerate(df):
    #Take out the trash
    dumpster = ['Memphis','usa',',','United States','Tennessee', ' tn ']
    zipFormat = r'\b38\d{3}\b' #Target zip codes

    def emptyDumpster(cell):
        if pd.isna(cell) or cell is None:
            return cell
        for trash in dumpster:
            cell = re.sub(trash, '', str(cell), flags=re.IGNORECASE)
        cell = re.sub(zipFormat, '', cell)
        return cell

    for col in df.columns:
        if col != 'Reported Date':
            df[col] = df[col].map(emptyDumpster)
    return df

def runGISGrabber():
    # Define the URL and parameters
    url = "https://maps.memphistn.gov/mapping/rest/services/PublicWorks/Drain_Services_PROD/FeatureServer/1/query"
    params = {
        "where": "1=1",
        "outFields": "*",
        "f": "json"
    }

    # Fetch data from the GIS server
    response = requests.get(url, params=params)

    if response.status_code == 200:
        data = response.json()

        if "features" in data and len(data["features"]) > 0:
            # Convert the feature data into a DataFrame
            df = pd.json_normalize(data["features"])

            # Define a manual mapping of old column names to aliases
            columnMapping = {
                'attributes.INCIDENT_NUMBER': 'Service Request Number',
                'attributes.REPORTED_DATE': 'Reported Date',
                'attributes.ADDRESS1': 'Location',
                'attributes.REQUEST_TYPE': 'Service Request Type ID',
                'attributes.REQUEST_SUMMARY': 'Service Request Summary',
                'attributes.Drain_Zone': 'Drain Zone',
                'attributes.MAP_PG': 'Map Page',
                'attributes.MAP_BLK': 'Map Block',
                'attributes.ASSIGNED_TO': 'Assigned To',
                'attributes.SCF_URL': 'SeeClickFix URL',
            }

            # Rename the columns based on the valid mappings
            df.rename(columns=columnMapping, inplace=True)

            # Remove any columns that are not in the columnMapping
            columnsToKeep = list(columnMapping.values())
            df = df[columnsToKeep]

            #Convert some dates
            df['Reported Date'] = pd.to_datetime(df['Reported Date'], unit='ms', origin='unix')

            #Sort based on Map Page
            df.sort_values(by='Map Page', ascending=True, inplace=True)
            #Delete unneeded nonsense
            incinerate(df)
            # Save to Excel
            currentDate = datetime.datetime.now().strftime('%Y-%m-%d__%H-%M-%S')
            outputPath = f'DrainDaily{currentDate}.xlsx'
            df.to_excel(outputPath, index=False)

            # Load the workbook to apply formatting
            wb = load_workbook(outputPath)
            ws = wb.active
            applyFormatting(df, ws, wb, outputPath)

            def showPrintMessage(message, title='Print?', style=0x4):
                return ctypes.windll.user32.MessageBoxW(0, message, title, style)
            
            def getPrintCopies():
                root = tk.Tk()
                root.withdraw()
                copies = simpledialog.askinteger('Print Copies','Enter the number of copies you want to print (max of 10), or enter 0 to cancel.', minvalue=0, maxvalue=10)
                return copies

            def printSheets(outputPath):
                response = showPrintMessage('Do you want to print the zone sheets?', 'Confirm Print', 0x4)
                #response = 6
                if response == 6:
                    copies = getPrintCopies()
                    if copies == 0 or None:
                        showPrintMessage('Printing Canceled','Info',0x40)
                    
                    excel = win32com.client.Dispatch('Excel.Application')
                    workbook = excel.Workbooks.Open(outputPath)

                    sheetsToPrint = [workbook.Sheets[i] for i in range(1,5)]
                    activePrinter = excel.ActivePrinter

                    for sheet in sheetsToPrint:
                        for _ in range(copies):
                            sheet.PrintOut(ActivePrinter = activePrinter, Copies=1, Collate=True)

                    workbook.Close(False)
                    excel.Quit()
                else:
                    showPrintMessage('Printing canceled, but file was still saved', 'Info')

            outputPath = os.path.abspath(f'DrainDaily{currentDate}.xlsx')
            printSheets(outputPath)
        else:
            print("No features found in the dataset.")
    else:
        def showErrorMessage():
            ctypes.windll.user32.MessageBoxW(0,
                                            "Error: Unable to retrieve data. Either the GIS site is down, you are not connected to the internet, or something else broke. Please contact Brian.",
                                            "Error",
                                            0x10)
        showErrorMessage()
if __name__ == '__main__':
    runGISGrabber()
