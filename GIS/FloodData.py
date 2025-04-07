import tkinter as tk
import pandas as pd
import requests
import datetime
from datetime import datetime
import re
import ctypes
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.table import Table 
from openpyxl.styles.fills import GradientFill
from openpyxl.worksheet.page import PageMargins
import openpyxl.styles
import win32com.client
from tkinter import simpledialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, GradientFill, Color
from openpyxl.worksheet.table import Table

def showErrorMessage():
    ctypes.windll.user32.MessageBoxW(0,
    "Error: Unable to retrieve data. Either GIS is down, you are not connected to the internet, or something else broke. Please contact Brian.",
    "Error",
    0x10)

def applyFormatting(df, ws, wb, outputPath):
    ws.title = 'All Drain Zones'

    # Set the format for the header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B7B7B7", end_color="B7B7B7", fill_type="solid")

    # Set column widths based on header content
    column_widths = {
        'A': 12.71,  # Service Request Number
        'B': 14.71,  # Reported Date
        'C': 23.71,  # Location
        'D': 22.71,  # Service Request Type ID
        'E': 61.71,  # Service Request Summary
        'F': 5.71,  # Drain Zone
        'G': 5.71,  # Map Page
        'H': 5.71,  # Map Block
        'I': 13.71,  # Assigned To
        'J': 0.71   # SeeClickFix URL does not matter since it gets hidden later
    }

    # Apply the column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.column_dimensions['J'].hidden = True #Hide J

    # Apply borders around the data cells
    from openpyxl.styles.borders import Border, Side
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    #Specific cell formats ORDER MATTERS and yes they have to be separate, no idk why

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for cell in ['A1','F1', 'G1', 'H1']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    for cell in ws['C']:
        cell.alignment = Alignment(horizontal='center', vertical='center',shrink_to_fit=True)

    for cell in ws['D']:
        cell.alignment = Alignment(horizontal='center', vertical='center',shrink_to_fit=True)

    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    for row in range(2, ws.max_row + 1): #Format col a based on j
        if ws[f'J{row}'].value:
            cell = ws[f'A{row}']

            fill = GradientFill(stop=("FF0000", "FFFFFF"))
            cell.fill = fill

    #Wrap the entire thing in a precious table
    table = Table(displayName="DataTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
    ws.add_table(table)

    # Apply misc formatting
    colBRange = f'B2:B{ws.max_row}'
    for cell in ws['B']:
        cell.number_format = 'm/d/yyyy'  # Apply short date format
    color_scale_rule = ColorScaleRule(
        start_type="min", start_color=Color(rgb="FFFF0000"),  # Red
        mid_type="percentile", mid_value=50, mid_color=Color(rgb="FFFFFF00"),  # Yellow
        end_type="max", end_color=Color(rgb="FF00FF00")  # Green
    )

    colDRange = f'D2:D{ws.max_row}'
    conditions = [
        ("flood", "FFC7CE"), #light peach
        ("prevent","FFC7C3"),
        ("repair","FFEB9C"), #light orange
        ("reset","FFEB9C"),
        ("cavity","D6DCE4") #light gray
    ]

    def applyConditionalFormatting(ws):
        ws.conditional_formatting.add(colBRange, color_scale_rule)
        for keyword, color in conditions:
            ws.conditional_formatting.add(
                colDRange,
                FormulaRule(formula=[f'ISNUMBER(SEARCH("{keyword}",D2))'], stopIfTrue=False,
                            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
            )
    applyConditionalFormatting(ws)

    zoneColors = {'A': '00FF00','B': 'FFFF00','C': '0000FF','D': 'FFA500'}

    #Copy the sheet, then color based on zone 
    for i in range(4):
        newSheet = wb.copy_worksheet(ws)
        zone = chr(65 + i)
        newSheet.title = f'Drain Zone {chr(65 + i)}'

        header_fill = PatternFill(start_color=zoneColors[zone], end_color=zoneColors[zone], fill_type="solid")
        for cell in newSheet[1]:  # Loop through all cells in row 1
            cell.fill = header_fill

        if zone == 'C':
            for cell in newSheet[1]:
                cell.font = Font(color='FFFFFF')

        applyConditionalFormatting(newSheet)

    #Pseudo Filter zones
    for i in range (2,6):
        sheet = wb[f'Drain Zone {chr(65 + i - 2)}']
        drainZone = chr(65 + i - 2)

        # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=6, max_col=6):
        #     if row[0].value != drainZone:
        #         sheet.row_dimensions[row[0].row].hidden = True

        #Thing above will hide the lines, one below will delete them. One below better bc color scale rule

        for rowNum in range(sheet.max_row, 1, -1):  #reverse to deal with any indexing problems
            if sheet[f'F{rowNum}'].value != drainZone: 
                sheet.delete_rows(rowNum)


    #Let's get this stupid thing ready to print
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Landscape orientation
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL  # Legal paper size
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.3, footer=0.3)
        ws.oddFooter.right.text = "&\"-,Bold\"&14&KFF0000&D"  # Date in red
        ws.oddFooter.left.text = "&\"-,Bold\"&14&KFF0000&P"  # Page number in the left footer in red


    # Save the workbook with formatting applied
    wb.save(outputPath)
    print(f"Data saved and formatted successfully to {outputPath}")

def incinerate(df):
    #Take out the trash
    dumpster = ['Memphis','usa',',','United States','Tennessee', ' tn ']
    zipFormat = r'\b38\d{3}\b' #Target zip codes

    def emptyDumpster(cell):
        if pd.isna(cell) or cell is None:
            return cell
        for trash in dumpster:
            cell = re.sub(trash, '', str(cell), flags=re.IGNORECASE)
        cell = re.sub(zipFormat, '', cell)
        return cell

    for col in df.columns:
        if col != 'Reported Date':
            df[col] = df[col].map(emptyDumpster)
    return df

def fetchFloodData(filter_condition):
    url = "https://maps.memphistn.gov/mapping/rest/services/PublicWorks/Drain_Services_PROD/FeatureServer/0/query"
    params = {
        "where": filter_condition,
        "outFields": "*",
        "f": "json"
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            # Check if "features" exist and have data
            if "features" in data and len(data["features"]) > 0:
                # Convert the feature data into a DataFrame
                df = pd.json_normalize(data["features"])

                # Define a manual mapping of old column names to aliases
                columnMapping = {
                    'attributes.INCIDENT_NUMBER': 'Service Request Number',
                    'attributes.REPORTED_DATE': 'Reported Date',
                    'attributes.ADDRESS1': 'Location',
                    'attributes.REQUEST_TYPE':'Service Request Type ID',
                    'attributes.REQUEST_SUMMARY':'Service Request Summary',
                    'attributes.Drain_Zone':'Drain Zone',
                    'attributes.MAP_PG':'Map Page',
                    'attributes.MAP_BLK':'Map Block',
                    'attributes.ASSIGNED_TO':'Assigned To',
                    'attributes.SCF_URL':'SeeClickFix URL'
                }

                # Rename the columns based on the manual mapping
                df.rename(columns=columnMapping, inplace=True)

                # Keep only the needed columns
                columnsToKeep = list(columnMapping.values())
                df = df[columnsToKeep]

                #Convert some dates
                df['Reported Date'] = pd.to_datetime(df['Reported Date'], unit='ms', origin='unix')
                #Sort by map page
                df.sort_values(by='Map Page', ascending=True, inplace=True)

                #Delete the nonsense
                incinerate(df)
                return df
            else:
                print("No features found in the data.")
                return pd.DataFrame()
        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")
            return pd.DataFrame()
    else:
        showErrorMessage()

def exportFloodData(df):
    if df.empty:
        print("No data to export.")
        return

    currentDate = datetime.now().strftime('%Y-%m-%d__%H-%M-%S')
    outputPath = f'FloodData{currentDate}.xlsx'

    # Create a workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'All Drain Zones'

    # Write DataFrame to worksheet
    for r_idx, row in enumerate(df.columns.to_list(), start=1):
        ws.cell(row=1, column=r_idx, value=row)

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    applyFormatting(df, ws, wb, outputPath)

    # Save the workbook
    def showPrintMessage(message, title='Print?', style=0x4):
        return ctypes.windll.user32.MessageBoxW(0, message, title, style)

    wb.save(outputPath)
    print(f"Data exported successfully to {outputPath}.")
    showPrintMessage(f'Exported successfully to {outputPath}','Info',0x40)

def queryFloodData():
    global exportButton, df3
    dateStr = simpledialog.askstring('Enter Date','Enter start date from which to pull flood data (mm/dd/yyyy).')
    if not dateStr:
        dateObj = datetime.now()
    else:
        try:
            dateObj = datetime.strptime(dateStr, "%m/%d/%Y")
        except ValueError:
            dateObj = datetime.now()
    sqlDate = dateObj.strftime("%Y-%m-%d")
    df1 = fetchFloodData(f"(REQUEST_TYPE LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%stand%' Or REQUEST_SUMMARY LIKE '%clog%' Or REQUEST_TYPE LIKE '%clean inlet%' Or REQUEST_SUMMARY LIKE '%clean inlet%') AND REQUEST_SUMMARY NOT LIKE '%hotspot%' And REPORTED_DATE >= timestamp '{sqlDate} 00:00:00'")
    #df2 = fetchFloodData(f"(REQUEST_TYPE LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%stand%' Or REQUEST_SUMMARY LIKE '%clog%' Or REQUEST_TYPE LIKE '%clean inlet%' Or REQUEST_SUMMARY LIKE '%clean inlet%') AND REQUEST_SUMMARY NOT LIKE '%hotspot%' And CLOSE_DATE >= timestamp '{sqlDate} 00:00:00'")
    df3 = fetchFloodData(f"(REQUEST_TYPE LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%flood%' Or REQUEST_SUMMARY LIKE '%stand%' Or REQUEST_SUMMARY LIKE '%clog%' Or REQUEST_TYPE LIKE '%clean inlet%' Or REQUEST_SUMMARY LIKE '%clean inlet%') AND REQUEST_SUMMARY NOT LIKE '%hotspot%' And REPORTED_DATE >= timestamp '{sqlDate} 00:00:00' And REQUEST_STATUS <> 'Closed'")
    df4 = fetchFloodData("REQUEST_SUMMARY LIKE '%hotspot%' And REQUEST_STATUS = 'Open'")
    df5 = fetchFloodData(f"REQUEST_SUMMARY LIKE '%hotspot%' And CLOSE_DATE >= timestamp '{sqlDate} 00:00:00'")
    
    floodingRequests = len(df1) if not df1.empty else 0
    floodingCompleted = len(df1) - len(df3) #Found through practice that they constantly close tickets wrong and mess up this number. So just using math. 
    outstandingFloods = len(df3) if not df3.empty else 0
    outstandingHotspots = len(df4) if not df4.empty else 0
    hotspotsCompleted = len(df5) if not df5.empty else 0
    
    resultLabel.config(text=(f"Flooding Requests Received since {sqlDate}: {floodingRequests}\n"
                            f"Flooding Requests Completed since {sqlDate}: {floodingCompleted}\n"
                            f"Outstanding Flooding Requests: {outstandingFloods}\n"
                            f"Hotspot Locations Completed since {sqlDate}: {hotspotsCompleted}\n"
                            f"Outstanding Hotspot Locations: {outstandingHotspots}"))

    # Show the "Export" button after fetching data
    exportButton.pack(pady=5)

def main():
    global resultLabel, exportButton
    root = tk.Tk()
    root.title("Ticket Grabber")
    root.geometry("500x300")
    submitButton = tk.Button(root, text="Retrieve Flood Data", command=queryFloodData)
    submitButton.pack(pady=5)

    resultLabel = tk.Label(root, text="")
    resultLabel.pack(pady=10)

    exportButton = tk.Button(root, text="Export Outstanding Flooding Tickets to Excel", command=lambda: exportFloodData(df3))
    exportButton.pack_forget()

    versionLabel = tk.Label(root, text="Version 4.7.25")
    versionLabel.place(x=10, y=275)

    root.mainloop()

currentDate = datetime.now()
checkDate = datetime(2026, 6, 1)
#Forcing myself to look through this every now and then 

if __name__ == "__main__" and checkDate > currentDate:
    main()
else:
    showErrorMessage()