import openpyxl.styles
import re
import requests
import datetime
import pandas as pd
import openpyxl
import win32com.client
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import os
import win32com.client
import tkinter as tk
from tkinter import filedialog, messagebox

# Define the URL and parameters
url = "https://maps.memphistn.gov/mapping/rest/services/PublicWorks/Drain_Services_PROD/FeatureServer/1/query"
params = {
    "where": "1=1",
    "outFields": "*",
    "f": "json"
}

# Fetch data from the GIS server
response = requests.get(url, params=params)

if response.status_code == 200:
    data = response.json()

    if "features" in data and len(data["features"]) > 0:
        # Convert the feature data into a DataFrame
        df = pd.json_normalize(data["features"])

        # Define a manual mapping of old column names to aliases
        columnMapping = {
            'attributes.INCIDENT_NUMBER': 'Service Request Number',
            'attributes.REPORTED_DATE': 'Reported Date',
            'attributes.ADDRESS1': 'Location',
            'attributes.REQUEST_TYPE': 'Service Request Type ID',
            'attributes.REQUEST_SUMMARY': 'Service Request Summary',
            'attributes.Drain_Zone': 'Drain Zone',
            'attributes.MAP_PG': 'Map Page',
            'attributes.MAP_BLK': 'Map Block',
            'attributes.ASSIGNED_TO': 'Assigned To',
            'attributes.SCF_URL': 'SeeClickFix URL',
        }

        # Rename the columns based on the valid mappings
        df.rename(columns=columnMapping, inplace=True)

        # Remove any columns that are not in the columnMapping
        columnsToKeep = list(columnMapping.values())
        df = df[columnsToKeep]

        #Convert some dates
        df['Reported Date'] = pd.to_datetime(df['Reported Date'], unit='ms', origin='unix')

        #Take out the trash
        dumpster = ['Memphis','usa',',','United States','Tennessee']
        zipFormat = r'\b38\d{3}\b' #Target zip codes

        def emptyDumpster(cell):
            if pd.isna(cell) or cell is None:
                return cell
            for trash in dumpster:
                cell = re.sub(trash, '', str(cell), flags=re.IGNORECASE)
            cell = re.sub(zipFormat, '', cell)
            return cell
        
        for col in df.columns:
            if col != 'Reported Date':
                df[col] = df[col].map(emptyDumpster)

        #Sort based on Map Page
        df.sort_values(by='Map Page', ascending=True, inplace=True)

        # Save to Excel
        currentDate = datetime.datetime.now().strftime('%Y-%m-%d')
        outputPath = f"C:\\Users\\brian.stlouis\\Documents\\DrainDaily{currentDate}.xlsx"
        df.to_excel(outputPath, index=False)

        # Load the workbook to apply formatting
        wb = load_workbook(outputPath)
        ws = wb.active

        # Apply formatting##########################################################

        # Set the format for the header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="B7B7B7", end_color="B7B7B7", fill_type="solid")

        # Set column widths based on header content
        column_widths = {
            'A': 12.71,  # Service Request Number
            'B': 14.71,  # Reported Date
            'C': 23.71,  # Location
            'D': 22.71,  # Service Request Type ID
            'E': 61.71,  # Service Request Summary
            'F': 5.71,  # Drain Zone
            'G': 5.71,  # Map Page
            'H': 5.71,  # Map Block
            'I': 13.71,  # Assigned To
            'J': 0.71   # SeeClickFix URL does not matter since it gets hidden later
        }

        # Apply the column widths
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.column_dimensions['J'].hidden = True #Hide J

        # Apply misc formatting

        # Applying a 3-color scale to column B
        col = 'B2:B' + str(ws.max_row)
        for cell in ws['B']:
            cell.number_format = 'm/d/yyyy'  # Apply short date format

        # Create a color scale rule (Red, Yellow, Green)
        color_scale_rule = ColorScaleRule(
            start_type="min", start_color=Color(rgb="FFFF0000"),  # Red
            mid_type="percentile", mid_value=50, mid_color=Color(rgb="FFFFFF00"),  # Yellow
            end_type="max", end_color=Color(rgb="FF00FF00")  # Green
        )

        # Apply the color scale rule
        ws.conditional_formatting.add(col, color_scale_rule)

        # Define formatting styles column D
        flood_prevent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light peach
        flood_prevent_font = Font(color='9C0006')

        repair_reset_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light orange
        repair_reset_font = Font(color='9C5700')

        cavity_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # Light gray

        # Apply formatting rules to column D (assume column index = 4)
        col_d_range = "D2:D" + str(ws.max_row)  # Adjusted to cover all rows

        # Conditional formatting rules col D
        ws.conditional_formatting.add(col_d_range, 
            FormulaRule(formula=['ISNUMBER(SEARCH("flood",D2))'], stopIfTrue=False, font=flood_prevent_font, fill=flood_prevent_fill))

        ws.conditional_formatting.add(col_d_range, 
            FormulaRule(formula=['ISNUMBER(SEARCH("prevent",D2))'], stopIfTrue=False, font=flood_prevent_font, fill=flood_prevent_fill))

        ws.conditional_formatting.add(col_d_range, 
            FormulaRule(formula=['ISNUMBER(SEARCH("repair",D2))'], stopIfTrue=False, font=repair_reset_font, fill=repair_reset_fill))

        ws.conditional_formatting.add(col_d_range, 
            FormulaRule(formula=['ISNUMBER(SEARCH("reset",D2))'], stopIfTrue=False, font=repair_reset_font, fill=repair_reset_fill))

        ws.conditional_formatting.add(col_d_range, 
            FormulaRule(formula=['ISNUMBER(SEARCH("cavity",D2))'], stopIfTrue=False, fill=cavity_fill))

        # Apply borders around the data cells
        from openpyxl.styles.borders import Border, Side
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        #Specific cell formats ORDER MATTERS and yes they have to be separate, no idk why

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        for cell in ['A1','F1', 'G1', 'H1']:
            ws[cell].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        for cell in ws['D']:
            cell.alignment = Alignment(horizontal='center', vertical='center',shrink_to_fit=True)

        for cell in ws['E']:
            cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        for row in range(2, ws.max_row + 1): #Format col a based on j
            if ws[f'J{row}'].value:
                cell = ws[f'A{row}']

                fill = GradientFill(stop=("FF0000", "FFFFFF"))
                cell.fill = fill

        #Let's get this thing ready to print 
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # Landscape orientation
        ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL  # Legal paper size
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.3, footer=0.3)
        ws.oddHeader.center.text = "&\"-,Bold\"&36&KFF0000" + "\n" + "&D"  # Date in the center header in red
        ws.oddFooter.left.text = "&\"-,Bold\"&14&KFF0000&P"  # Page number in the left footer in red

        #Wrap the entire thing in a precious table
        table = Table(displayName="DataTable", ref=f"A1:{chr(65 + len(df.columns) - 1)}{len(df) + 1}")
        ws.add_table(table)

        # Save the workbook with formatting applied
        wb.save(outputPath)
        print(f"Data saved and formatted successfully to {outputPath}")

        #Print
        def printFile(outputPath):
            excel = win32com.client.Dispatch('Excel.Application')
            workbook = excel.Workbooks.Open(outputPath)
            workbook.PrintOut()
            workbook.Close(False)
            excel.Quit()
        
        #printFile(outputPath)

    else:
        print("No features found in the dataset.")
else:
    print(f"Error: {response.status_code} - {response.text}")

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Drain Services Data Export")

        # Default values
        self.output_path = ""
        self.selected_zones = []
        self.print_selected = False
        self.num_copies = 1

        # Create UI elements
        self.create_widgets()

    def create_widgets(self):
        # Label for output path
        self.output_label = tk.Label(self.root, text="Select Output Path:")
        self.output_label.grid(row=0, column=0, padx=10, pady=10)

        # Button to select output path
        self.select_output_button = tk.Button(self.root, text="Select Folder", command=self.select_output_path)
        self.select_output_button.grid(row=0, column=1, padx=10, pady=10)

        # Label for Drain Zone options
        self.drain_zone_label = tk.Label(self.root, text="Select Drain Zones:")
        self.drain_zone_label.grid(row=1, column=0, padx=10, pady=10)

        # Checkboxes for A, B, C, D zones
        self.zone_a_var = tk.BooleanVar()
        self.zone_b_var = tk.BooleanVar()
        self.zone_c_var = tk.BooleanVar()
        self.zone_d_var = tk.BooleanVar()

        self.zone_a_checkbox = tk.Checkbutton(self.root, text="A", variable=self.zone_a_var)
        self.zone_a_checkbox.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        self.zone_b_checkbox = tk.Checkbutton(self.root, text="B", variable=self.zone_b_var)
        self.zone_b_checkbox.grid(row=1, column=2, padx=10, pady=5, sticky="w")

        self.zone_c_checkbox = tk.Checkbutton(self.root, text="C", variable=self.zone_c_var)
        self.zone_c_checkbox.grid(row=1, column=3, padx=10, pady=5, sticky="w")

        self.zone_d_checkbox = tk.Checkbutton(self.root, text="D", variable=self.zone_d_var)
        self.zone_d_checkbox.grid(row=1, column=4, padx=10, pady=5, sticky="w")

        # Print options
        self.print_label = tk.Label(self.root, text="Print Options:")
        self.print_label.grid(row=2, column=0, padx=10, pady=10)

        self.print_var = tk.BooleanVar()
        self.print_checkbox = tk.Checkbutton(self.root, text="Print", variable=self.print_var, command=self.toggle_print)
        self.print_checkbox.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        self.print_count_label = tk.Label(self.root, text="Number of Copies:")
        self.print_count_label.grid(row=3, column=0, padx=10, pady=5)

        self.print_count_entry = tk.Entry(self.root)
        self.print_count_entry.grid(row=3, column=1, padx=10, pady=5)
        self.print_count_entry.insert(0, "1")  # Default to 1 copy

        # Export button
        self.export_button = tk.Button(self.root, text="Export and Print", command=self.export_data)
        self.export_button.grid(row=4, column=0, columnspan=2, padx=10, pady=20)

    def select_output_path(self):
        self.output_path = filedialog.askdirectory(title="Select Output Folder")
        if self.output_path:
            messagebox.showinfo("Selected Path", f"Output will be saved to:\n{self.output_path}")

    def toggle_print(self):
        if self.print_var.get():
            self.print_count_label.grid(row=3, column=0, padx=10, pady=5)
            self.print_count_entry.grid(row=3, column=1, padx=10, pady=5)
        else:
            self.print_count_label.grid_forget()
            self.print_count_entry.grid_forget()

    def export_data(self):
        if not self.output_path:
            messagebox.showerror("Error", "Please select an output path.")
            return

        # Collect selected zones
        self.selected_zones = []
        if self.zone_a_var.get():
            self.selected_zones.append("A")
        if self.zone_b_var.get():
            self.selected_zones.append("B")
        if self.zone_c_var.get():
            self.selected_zones.append("C")
        if self.zone_d_var.get():
            self.selected_zones.append("D")

        if not self.selected_zones:
            messagebox.showerror("Error", "Please select at least one Drain Zone.")
            return

        # Load the data (modify this based on your dataframe)
        df = pd.read_excel(outputPath)

        # Filter by selected zones
        filtered_df = df[df['Drain Zone'].isin(self.selected_zones)]

        # Save the filtered data
        file_name = f"Drain_Services_Data_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
        file_path = os.path.join(self.output_path, file_name)
        filtered_df.to_excel(file_path, index=False)

        # Optionally print the file
        if self.print_var.get():
            try:
                num_copies = int(self.print_count_entry.get())
                self.print_excel_file(file_path, num_copies)
                messagebox.showinfo("Success", f"File saved and sent to print.\n{num_copies} copies.")
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number of copies.")
        else:
            messagebox.showinfo("Success", f"File saved to:\n{file_path}")

    def print_excel_file(self, file_path, num_copies):
        # Open the Excel application
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False

        # Open the workbook
        workbook = excel.Workbooks.Open(file_path)

        # Print the workbook the specified number of times
        for _ in range(num_copies):
            workbook.PrintOut()

        # Close the workbook
        workbook.Close(False)

        # Quit Excel
        excel.Quit()

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
